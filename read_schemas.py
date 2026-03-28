import openpyxl
import os

data_dir = "data"
files = sorted([f for f in os.listdir(data_dir) if f.endswith('.xlsx')])

for f in files:
    wb = openpyxl.load_workbook(os.path.join(data_dir, f), read_only=True)
    for s in wb.sheetnames:
        rows = list(wb[s].iter_rows(min_row=1, max_row=1))
        cols = [c.value for c in rows[0]] if rows else []
        max_row = wb[s].max_row
        print(f"FILE: {f}")
        print(f"  SHEET: {s}")
        print(f"  ROWS: {max_row}")
        print(f"  COLS: {cols}")
        print()
    wb.close()
